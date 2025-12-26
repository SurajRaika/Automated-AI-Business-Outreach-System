"""Excel file management for scraping data, tracking email status within the spreadsheet."""

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from config import EXCEL_HEADERS, TABLE_STYLE, EXCEL_DIRECTORY # Original imports
import logging
import os
import re
import datetime



EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[a-zA-Z0-9]+$")
SENT_STATUS_COL_IDX = 9   
SENT_DATE_COL_IDX = 10 

class ExcelManager:
    """Manages Excel file operations for storing scraped data"""
    
    def __init__(self, filename):
        # Ensure directory exists
        if not os.path.exists(EXCEL_DIRECTORY):
            os.makedirs(EXCEL_DIRECTORY)
        
        # Add directory to filename if not already there
        if not filename.startswith(EXCEL_DIRECTORY):
            self.filename = os.path.join(EXCEL_DIRECTORY, filename)
        else:
            self.filename = filename
        
        self.sheet_name = 'Companies'
        self.headers = EXCEL_HEADERS
        self.table_style = TABLE_STYLE
        
        logging.info(f"Initializing ExcelManager for {self.filename}")
        self.wb = None  # Initialize as None to prevent AttributeErrors
        self._create_or_load_workbook()

    def insert_data(self, company_name, company_website, company_email, 
                   company_phone, rating_ratio, number_of_ratings, 
                   social_media_accounts='', city=''):
        """Insert new company data into the sheet"""
        if not company_name or not company_name.strip():
            logging.warning("Attempted to insert data with empty company name")
            return False
        
        # CHECK DUPLICATES (Matches Website + City only)
        if self.is_data_present(company_phone, company_website, city):
            logging.info(f"Data already present for {company_name} in {city}")
            return False

        row_data = [
            company_name, 
            company_website, 
            company_email, 
            company_phone, 
            rating_ratio, 
            number_of_ratings, 
            social_media_accounts,
            city,
            'False', # Sent Status
            ''       # Sent Date
        ]
        
        self.ws.append(row_data)
        self.wb.save(self.filename)
        logging.info(f"Inserted new data: {company_name} - {city}")
        return True
          
    def _create_or_load_workbook(self):
        """Create new workbook or load existing one - FIXED CRASH LOGIC"""
        try:
            self.wb = load_workbook(self.filename)
            logging.info(f"Loaded existing workbook: {self.filename}")
        except FileNotFoundError:
            self.wb = Workbook()
            logging.info(f"Created new workbook: {self.filename}")
        except Exception as e:
            logging.error(f"Error loading workbook (possibly corrupted): {e}. Creating new one.")
            self.wb = Workbook()

        # Ensure sheet exists (Moved out of finally block to avoid scope issues)
        if self.sheet_name in self.wb.sheetnames:
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb.active
        
        self._create_or_load_sheet()
        self.wb.save(self.filename)

    def _create_or_load_sheet(self):
        """Create sheet with headers if it doesn't exist or verify existing headers"""
        if self.ws.max_row == 1 and not self.ws['A1'].value:
            self.ws.title = self.sheet_name
            self._add_headers()
            self._add_table()
        elif self.ws.title != self.sheet_name:
            self.ws.title = self.sheet_name
            if self.ws.max_row <= 1 or not self.ws['A1'].value:
                 self._add_headers()
                 self._add_table()
            else:
                 self._update_table_range()
        else:
            self._update_table_range()

    def _add_headers(self):
        for col_idx, header in enumerate(self.headers, start=1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

    def _add_table(self):
        last_col = get_column_letter(len(self.headers))
        table_ref = f"A1:{last_col}{self.ws.max_row}"
        table = Table(displayName="Table1", ref=table_ref)
        style = TableStyleInfo(name=self.table_style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        self.ws.add_table(table)

    def _update_table_range(self):
        last_col = get_column_letter(len(self.headers))
        table_ref = f"A1:{last_col}{self.ws.max_row}"
        if self.ws.tables and 'Table1' in self.ws.tables:
            table = self.ws.tables['Table1']
            if table.ref != table_ref:
                table.ref = table_ref
        else:
             self._add_table()

    def is_data_present(self, company_phone, company_website, city):
        """
        Check if data exists based ONLY on Website and City.
        Returns False if Website or City is missing to allow insertion
        """
        if not company_website or not city:
            return True

        for row in self.ws.iter_rows(min_row=2, values_only=True):
            row_website = row[1]
            row_city = row[7]
            
            if row_website == company_website and row_city == city:
                return True

        return False
    
    def update_row_with_details(self, company_name, company_website, company_email, 
                                company_phone, rating_ratio, number_of_ratings, 
                                social_media_accounts, city):
        """Update row if exists, otherwise insert."""
        updated = False
        
        for row in self.ws.iter_rows(min_row=2):
            if row[1].value == company_website and row[7].value == city:
                row[2].value = company_email if company_email else 'None'
                
                if not row[3].value or row[3].value == 'None':
                     row[3].value = company_phone if company_phone else 'None'
                if not row[4].value or row[4].value == 'None':
                    row[4].value = rating_ratio if rating_ratio else 'None'
                if not row[5].value or row[5].value == 'None':
                    row[5].value = number_of_ratings if number_of_ratings else 'None'
                
                if isinstance(social_media_accounts, list):
                    row[6].value = ', '.join(social_media_accounts) if social_media_accounts else 'None'
                else:
                    row[6].value = social_media_accounts if social_media_accounts else 'None'
                
                self.wb.save(self.filename)
                updated = True
                break
        
        if not updated:
            return self.insert_data(company_name, company_website, company_email, 
                                    company_phone, rating_ratio, number_of_ratings, 
                                    social_media_accounts, city)
        return updated

    # --- HELPER METHODS (Abbreviated for space, they remain largely the same) ---
    
    def flag_company_as_sent(self, website_link, email_address, sent_status):
        found_and_updated = False
        timestamp = datetime.datetime.now().isoformat(timespec='minutes') if sent_status else ""
        status_str = "True" if sent_status else "False"

        for row in self.ws.iter_rows(min_row=2):
            if row[1].value == website_link and row[2].value == email_address:
                row[SENT_STATUS_COL_IDX - 1].value = status_str
                row[SENT_DATE_COL_IDX - 1].value = timestamp
                found_and_updated = True
                break

        if found_and_updated:
            self.wb.save(self.filename)
            return True
        return False

    def _email_already_sent(self, email_address):
        target_email = email_address.strip()
        for row in self.ws.iter_rows(min_row=2):
            row_email = row[2].value
            if row_email and row_email.strip() == target_email:
                sent_status_value = row[SENT_STATUS_COL_IDX - 1].value
                if isinstance(sent_status_value, str) and sent_status_value.strip().lower() == 'true':
                    return True
        return False

    def has_contact_info(self, company_website, city):
        for row in self.ws.iter_rows(min_row=2):
            if row[1].value == company_website and row[7].value == city:
                emails = row[2].value
                if emails and isinstance(emails, str) and emails.strip().lower() not in ("", "none"):
                    return True
        return False


    def delete_row_without_email(self, company_website, city):
        """Delete row if no email was found"""
        rows_to_delete = []
        
        for idx, row in enumerate(self.ws.iter_rows(min_row=2), start=2):
            if row[1].value == company_website and row[7].value == city:
                emails = row[2].value
                # If email is None, empty, or "None" string
                if not emails or emails == "None":
                    rows_to_delete.append(idx)
        
        # Delete in reverse order to maintain indices
        for row_idx in reversed(rows_to_delete):
            self.ws.delete_rows(row_idx)
            logging.info(f"Deleted row {row_idx} - no email found for {company_website} in {city}")
        
        if rows_to_delete:
            self.wb.save(self.filename)
            return True
        return False

    def get_companies_without_emails(self, city=None):
        """Get list of companies that need email extraction"""
        companies = []
        
        for row in self.ws.iter_rows(min_row=2):
            website = row[1].value
            email = row[2].value
            row_city = row[7].value
            company_name = row[0].value
            
            # Filter by city if specified
            if city and row_city != city:
                continue
            
            # Check if needs email extraction
            if not email or email == "None":
                companies.append({
                    'name': company_name,
                    'website': website,
                    'city': row_city
                })
        
        return companies
        
    def get_companies_with_emails(self, city=None):
        """Get list of companies that already have an email recorded"""
        companies = []
        
        for row in self.ws.iter_rows(min_row=2):
            website = row[1].value
            email = row[2].value
            row_city = row[7].value
            company_name = row[0].value
            
            # Filter by city if specified
            if city and row_city != city:
                continue
            
            # Include only records with an email
            if email and email != "None":
                companies.append({
                    'name': company_name,
                    'website': website,
                    'email': email,
                    'city': row_city
                })
        
        return companies

    def get_row_count(self):
        """Get total number of data rows (excluding header)"""
        return self.ws.max_row - 1

    def close(self):
        try:
            self.wb.save(self.filename)
            self.wb.close()
        except Exception as e:
            logging.error(f"Error closing workbook: {e}")